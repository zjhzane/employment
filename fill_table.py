# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook

# ===== 桌面路径与文件 =====
DESKTOP = Path.home() / "Desktop"
SRC  = DESKTOP / "毕业派遣管理_2025-09-01_19-40-17_595.xlsx"   # ← 按需改为你的源文件名
TMPL = DESKTOP / "new.xlsx"
OUT  = DESKTOP / "new_filled.xlsx"
# ========================

# 源表字段别名：尽量兼容不同列名（可按需补充）
FIELD_ALIASES = {
    "姓名": ["姓名","学生姓名","人员姓名","名字"],
    "生源地名称": ["生源地名称","生源地","生源所在地","户籍所在地","户籍"],
    "档案转寄类型名称": ["档案转寄类型名称","档案转寄类型","档案转递类型","转寄类型","转递类型"],
    "身份证号": ["身份证号","身份证号码","公民身份号码","证件号码"],
    "手机号码": ["手机号码","手机号","联系电话","联系手机","手机"],
    "用人单位名称": ["用人单位名称","用人单位","单位名称","就业单位名称","接收单位名称"],
    "档案转寄单位": ["档案转寄单位","档案转递单位","档案接收单位","接收单位","档案邮寄单位"],
    "档案转递单位地址": ["档案转递单位地址","档案接收单位地址","接收单位地址","档案转寄单位地址","转寄单位地址","邮寄地址","地址","家庭地址"],
    "档案转寄联系人": ["档案转寄联系人","接收单位联系人","联系人"],
    "档案转寄联系电话": ["档案转寄联系电话","接收单位联系电话","联系人电话","联系电话","电话","手机"],
    "班级": ["班级","所属班级","班级名称"],
    # 明确不写入的字段
    "转递编号": ["转递编号","档案转递编号","编号"],
}

# 这些列写入时强制用“字符串”以保留前导零
FORCE_TEXT_FIELDS = {"身份证号", "手机号码", "档案转寄联系电话", "联系电话", "转递编号"}

def norm(s):
    return "" if s is None else str(s).strip().replace("\r","").replace("\n","")

def read_header_map(ws):
    """读取首行表头：表头名 -> 列号（1-based）"""
    h = {}
    for j, cell in enumerate(ws[1], start=1):
        name = norm(cell.value)
        if name and name not in h:
            h[name] = j
    return h

def find_source_col(template_header, src_header_map):
    """
    给定模板列名，在源表头中寻找最合适的列号：
    1) 模板列名同名命中
    2) 在别名列表中按“完全等于”匹配
    3) 在别名列表中按“包含关系”匹配
    找不到返回 None
    """
    th = template_header
    # 1) 同名
    if th in src_header_map:
        return src_header_map[th]
    # 2) 别名（完全等于）
    for alias in FIELD_ALIASES.get(th, []):
        if alias in src_header_map:
            return src_header_map[alias]
    # 3) 别名（包含）
    for alias in FIELD_ALIASES.get(th, []):
        for k in src_header_map.keys():
            if alias in k:
                return src_header_map[k]
    return None

def main():
    if not SRC.exists():
        raise FileNotFoundError(f"找不到源表：{SRC}")
    if not TMPL.exists():
        raise FileNotFoundError(f"找不到模板：{TMPL}")

    # 打开源表与模板
    wb_src  = load_workbook(SRC, data_only=True)
    ws_src  = wb_src.active
    wb_tmpl = load_workbook(TMPI:=TMPL)  # 兼容旧版本写法
    ws_tmpl = wb_tmpl.active

    # 表头映射
    src_h = read_header_map(ws_src)
    tmpl_h = read_header_map(ws_tmpl)
    if not tmpl_h:
        raise RuntimeError("模板 new.xlsx 首行表头为空，请先在第1行配置表头。")

    # 计算每个“模板列”在源表中的列号（None 表示留空）
    # 对“转递编号”强制置为 None（不填写）
    src_col_for_tmpl = {}
    for tmpl_col_name in [c for c in tmpl_h.keys()]:  # 遵循模板原始列顺序
        if tmpl_col_name == "转递编号":
            src_col_for_tmpl[tmpl_col_name] = None
        else:
            src_col_for_tmpl[tmpl_col_name] = find_source_col(tmpl_col_name, src_h)

    # 清空模板第2行开始的旧数据
    if ws_tmpl.max_row > 1:
        ws_tmpl.delete_rows(2, ws_tmpl.max_row - 1)

    # 源表从第2行开始，逐行拷贝到模板（严格按模板列顺序写）
    out_row = 2
    # 找到用于跳过空行的关键列（优先姓名，其次身份证）：
    key_name_col = find_source_col("姓名", src_h)
    key_id_col   = find_source_col("身份证号", src_h)

    for r in range(2, ws_src.max_row + 1):
        name_val = norm(ws_src.cell(row=r, column=key_name_col).value) if key_name_col else ""
        id_val   = norm(ws_src.cell(row=r, column=key_id_col).value)   if key_id_col else ""
        # 空记录跳过：姓名与身份证都空
        if not name_val and not id_val:
            continue

        # 按模板列顺序逐列写
        for tmpl_col_name, tmpl_col_idx in tmpl_h.items():
            src_col = src_col_for_tmpl.get(tmpl_col_name)
            if src_col is None:
                value = ""  # 留空（包含“转递编号”）
            else:
                raw = ws_src.cell(row=r, column=src_col).value
                value = norm(raw)
                if tmpl_col_name in FORCE_TEXT_FIELDS:
                    value = value  # 已是字符串
            ws_tmpl.cell(row=out_row, column=tmpl_col_idx, value=value)

        out_row += 1

    wb_tmpl.save(OUT)
    print(f"[OK] 已按模板列顺序填写完成，从第2行开始，共写入 {out_row-2} 行。")
    print(f"[OK] 输出文件：{OUT}")

if __name__ == "__main__":
    main()
